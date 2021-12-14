#!/usr/bin/env python
# _*_ coding:utf-8 _*_
import os
import sys

from src.core import settings

import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import Color, COLOR_INDEX

from src.core.settings import TESTER


class WriteExcel:
	"""文件写入数据"""

	def __init__(self, fileName):
		self.filename = fileName
		if not os.path.exists(self.filename):
			# 文件不存在，则拷贝模板文件至指定报告目录下
			shutil.copyfile(settings.SOURCE_FILE, settings.TARGET_FILE)
		self.wb = load_workbook(self.filename)
		self.ws = self.wb.active

	def write_data(self, row_n, value):
		"""
		写入测试结果
		:param row_n:数据所在行数
		:param value: 测试结果值
		:return: 无
		"""
		font_GREEN = Font(name='宋体', color=Color(rgb=COLOR_INDEX[3]), bold=True)
		font_RED = Font(name='宋体', color=Color(rgb=COLOR_INDEX[2]), bold=True)
		font1 = Font(name='宋体', color=Color(rgb=COLOR_INDEX[5]), bold=True)
		align = Alignment(horizontal='center', vertical='center')
		# 获数所在行数
		L_n = "L" + str(row_n)
		M_n = "M" + str(row_n)
		if value == "PASS":
			self.ws.cell(row_n, 12, value)
			self.ws[L_n].font = font_GREEN
		if value == "FAIL":
			self.ws.cell(row_n, 12, value)
			self.ws[L_n].font = font_RED
		self.ws.cell(row_n, 13, TESTER)
		self.ws[L_n].alignment = align
		self.ws[M_n].font = font1
		self.ws[M_n].alignment = align
		self.wb.save(self.filename)
