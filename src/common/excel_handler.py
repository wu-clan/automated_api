#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
import shutil

import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from src.common.log import log
from src.core.conf import settings
from src.core.path_settings import XLSX_FILE, EXCEL_REPORT, XLSX_FILE_TEMPLATE
from src.utils.time_control import get_current_time


def read_excel(filename: str, *, sheet_name: str = 'Sheet1') -> list:
    """
    读取excel文件

    :param filename:
    :param sheet_name:
    :return:
    """
    file = os.path.join(XLSX_FILE, filename)
    data = xlrd.open_workbook(file)
    table = data.sheet_by_name(sheet_name)
    # 获取总行数、总列数
    rows = table.nrows
    cols = table.ncols
    if rows > 1:
        # 获取第一行的内容，列表格式
        keys = table.row_values(0)
        data_list = []
        # 获取每一行的内容，列表格式
        for col in range(1, rows):
            values = table.row_values(col)
            # keys，values组合转换为字典
            api_dict = dict(zip(keys, values))
            data_list.append(api_dict)
        return data_list
    else:
        log.warning(f'数据表格 {filename} 没有数据!')
        raise ValueError(f'数据表格 {filename} 没有数据! 请检查数据文件内容是否正确!')


def write_excel(filename: str = f'APITestResult_{get_current_time()}.xlsx', *, row_num: int, status: str):
    """
    写入测试结果

    :param filename: 文件名
    :param row_num: 数据所在行数
    :param status: 测试结果值
    :return
    """
    status_upper = status.upper()
    if status_upper not in ('PASS', 'FAIL'):
        raise ValueError('excel测试报告结果用力状态只允许"PASS","FAIL')
    if not os.path.exists(EXCEL_REPORT):
        os.makedirs(EXCEL_REPORT)
    _filename = os.path.join(EXCEL_REPORT, filename)
    shutil.copyfile(XLSX_FILE_TEMPLATE, _filename)
    wb = load_workbook(_filename)
    ws = wb.active
    # 字体颜色
    font_green = Font(name='Consolas', color='99CC00', bold=True)
    font_red = Font(name='Consolas', color='FF0000', bold=True)
    font_black = Font(name='Consolas', color='000000', bold=True)
    align = Alignment(horizontal='center', vertical='center')
    # 获数所在行数
    l_num = "L" + str(row_num)
    m_num = "M" + str(row_num)
    # 写入测试结果
    if status_upper == "PASS":
        ws.cell(row_num, 12, status)
        ws[l_num].font = font_green
    if status_upper == "FAIL":
        ws.cell(row_num, 12, status)
        ws[l_num].font = font_red
    # 写入测试员
    ws.cell(row_num, 13, settings.TESTER_NAME)
    ws[m_num].font = font_black
    # 修改写入单元格样式
    ws[l_num].alignment = ws[m_num].alignment = align
    try:
        wb.save(_filename)
    except Exception as e:
        log.error(f'保存excel测试报告失败: {e}')
    else:
        log.info('写入excel测试报告成功')
