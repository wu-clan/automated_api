#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
import shutil
import time

import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Alignment
from openpyxl.styles.colors import COLOR_INDEX

from src.common.log import log
from src.core.conf import settings
from src.core.path_settings import XLSX_FILE, EXCEL_REPORT, TEMPLATE_XLSX_FILE

curr_time = time.strftime('%Y-%m-%d %H_%M_%S')


def read_excel(filename: str, sheetName: str = 'Sheet1'):
    """
    读取excel文件
    :param filename:
    :param sheetName:
    :return:
    """
    file = os.path.join(XLSX_FILE, filename)
    data = xlrd.open_workbook(file)
    table = data.sheet_by_name(sheetName)
    # 获取总行数、总列数
    rows = table.nrows
    cols = table.ncols
    if rows > 1:
        # 获取第一行的内容，列表格式
        keys = table.row_values(0)
        data_list = []
        # 获取每一行的内容，列表格式
        for col in range(1, rows):
            values = table.row_values(col)
            # keys，values组合转换为字典
            api_dict = dict(zip(keys, values))
            data_list.append(api_dict)
        return data_list
    else:
        log.warning("数据表格没有数据!")
        return None


def write_excel(row_n: int, status: str, filename: str = f'APITestResult_{curr_time}.xlsx'):
    """
    写入测试结果
    :param row_n:数据所在行数
    :param status: 测试结果值
    :param filename: 文件名
    :return
    """
    if not os.path.exists(EXCEL_REPORT):
        os.makedirs(EXCEL_REPORT)
    _filename = os.path.join(EXCEL_REPORT, filename)
    shutil.copyfile(TEMPLATE_XLSX_FILE, _filename)
    wb = load_workbook(_filename)
    ws = wb.active
    font_green = Font(name='宋体', color=Color(rgb=COLOR_INDEX[3]), bold=True)
    font_red = Font(name='宋体', color=Color(rgb=COLOR_INDEX[2]), bold=True)
    font_yellow = Font(name='宋体', color=Color(rgb=COLOR_INDEX[51]), bold=True)
    align = Alignment(horizontal='center', vertical='center')
    # 获数所在行数
    L_n = "L" + str(row_n)
    M_n = "M" + str(row_n)
    if status == "PASS":
        ws.cell(row_n, 12, status)
        ws[L_n].font = font_green
    if status == "FAIL":
        ws.cell(row_n, 12, status)
        ws[L_n].font = font_red
    ws.cell(row_n, 13, settings.TESTER_NAME)
    ws[M_n].font = font_yellow
    ws[L_n].alignment = ws[M_n].alignment = align
    try:
        wb.save(_filename)
    except Exception as e:
        log.error(f'保存excel测试报告失败\n{e}')
    else:
        log.info('写入excel测试报告成功')
