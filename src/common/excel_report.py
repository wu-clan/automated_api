#!/usr/bin/env python
# _*_ coding:utf-8 _*_
import os
import shutil

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import Color, COLOR_INDEX

from src.common.log import log
from src.core.conf import settings
from src.core.path_settings import EXCEL_RESULT, TEMPLATE_XLSX_FILE, EXCEL_REPORT


class WriteExcel:
	"""文件写入数据"""

	def __init__(self):
		if not os.path.exists(EXCEL_REPORT):
			os.makedirs(EXCEL_REPORT)
		if not os.path.exists(EXCEL_RESULT):
			shutil.copyfile(TEMPLATE_XLSX_FILE, EXCEL_RESULT)
		self.wb = load_workbook(EXCEL_RESULT)
		self.ws = self.wb.active

	def write_data(self, row_n, value):
		"""
		写入测试结果
		:param row_n:数据所在行数
		:param value: 测试结果值
		:return
		"""
		font_green = Font(name='宋体', color=Color(rgb=COLOR_INDEX[3]), bold=True)
		font_red = Font(name='宋体', color=Color(rgb=COLOR_INDEX[2]), bold=True)
		font_yellow = Font(name='宋体', color=Color(rgb=COLOR_INDEX[51]), bold=True)
		align = Alignment(horizontal='center', vertical='center')
		# 获数所在行数
		L_n = "L" + str(row_n)
		M_n = "M" + str(row_n)
		if value == "PASS":
			self.ws.cell(row_n, 12, value)
			self.ws[L_n].font = font_green
		if value == "FAIL":
			self.ws.cell(row_n, 12, value)
			self.ws[L_n].font = font_red
		self.ws.cell(row_n, 13, settings.TESTER_NAME)
		self.ws[M_n].font = font_yellow
		self.ws[L_n].alignment = self.ws[M_n].alignment = align
		try:
			self.wb.save(EXCEL_RESULT)
		except Exception as e:
			log.error(f'保存excel测试报告失败\n{e}')
		log.success('保存excel测试报告成功')
